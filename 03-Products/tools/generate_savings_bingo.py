import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_savings_bingo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savings Bingo"

    # --- Styles ---
    title_font = Font(name='Arial', size=24, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid") # Hot Pink

    cell_font = Font(name='Arial', size=16, bold=True)
    cell_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid") # Lavender

    center_align = Alignment(horizontal="center", vertical="center")
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Title ---
    ws.merge_cells('B2:F2')
    title_cell = ws['B2']
    title_cell.value = "Savings Bingo: 1,000 THB Challenge"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = thick_border

    # --- Bingo Grid (5x5) ---
    # Total sum should be 1000.
    # Let's distribute amounts: 20, 30, 40, 50, 60
    # 25 cells. Average cell value = 40.

    bingo_values = [
        20, 50, 30, 40, 20,
        40, 60, 20, 50, 30,
        30, 20, 100, 40, 50,  # Middle 100 is the "Jackpot"
        50, 40, 30, 60, 20,
        20, 30, 50, 40, 20
    ]
    # Sum check: 20*8 + 30*5 + 40*5 + 50*4 + 60*2 + 100*1 = 160 + 150 + 200 + 200 + 120 + 100 = 930 (Close enough, let's adjust to exactly 1000)
    # Adjusted: 20*5 + 30*5 + 40*5 + 50*5 + 60*5 = 100 + 150 + 200 + 250 + 300 = 1000! Perfect.

    bingo_values_balanced = [
        20, 30, 40, 50, 60,
        60, 50, 40, 30, 20,
        20, 30, 40, 50, 60,
        60, 50, 40, 30, 20,
        20, 30, 40, 50, 60
    ]

    start_row = 4
    start_col = 2

    idx = 0
    for r in range(5):
        for c in range(5):
            cell = ws.cell(row=start_row + r, column=start_col + c)
            cell.value = bingo_values_balanced[idx]
            cell.font = cell_font
            cell.fill = cell_fill
            cell.alignment = center_align
            cell.border = thin_border
            idx += 1

    # --- Formatting ---
    for col in range(start_col, start_col + 5):
        ws.column_dimensions[get_column_letter(col)].width = 15

    for row in range(start_row, start_row + 5):
        ws.row_dimensions[row].height = 40

    # --- Instructions ---
    ws.merge_cells('B10:F10')
    instr = ws['B10']
    instr.value = "Instructions: Save the amount shown, then cross it out (X)!"
    instr.alignment = center_align
    instr.font = Font(italic=True)

    # Save file
    output_path = "../templates/savings_bingo.xlsx"
    wb.save(output_path)
    print(f"Savings Bingo created successfully at: {output_path}")

if __name__ == "__main__":
    create_savings_bingo()
