import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_budget_template():
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Sheet 1: Dashboard (Summary) ---
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_dash['B2'] = "Monthly Budget Summary"
    ws_dash['B2'].font = Font(size=16, bold=True)

    summary_data = [
        ("Total Income", "=SUM(Income!C:C)"),
        ("Total Expenses", "=SUM(Expenses!C:C)"),
        ("Net Savings", "=B4-B5"),
        ("Savings Rate (%)", "=IF(B4=0, 0, B6/B4)")
    ]

    start_row = 4
    for i, (label, formula) in enumerate(summary_data):
        cell_label = ws_dash.cell(row=start_row + i, column=2, value=label)
        cell_value = ws_dash.cell(row=start_row + i, column=3, value=formula)

        cell_label.font = Font(bold=True)
        if label == "Net Savings":
            cell_label.font = Font(bold=True, color="006400") # Dark Green
            cell_value.font = Font(bold=True, color="006400")

        if label == "Savings Rate (%)":
             cell_value.number_format = '0.00%'
        else:
             cell_value.number_format = '#,##0.00'

    ws_dash.column_dimensions['B'].width = 20
    ws_dash.column_dimensions['C'].width = 15

    # --- Sheet 2: Income ---
    ws_income = wb.create_sheet("Income")
    headers_income = ["Date", "Source", "Amount", "Category", "Note"]

    for col_num, header in enumerate(headers_income, 1):
        cell = ws_income.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws_income.column_dimensions[get_column_letter(col_num)].width = 15

    # Add some dummy data
    dummy_income = [
        ("2023-10-01", "Salary", 25000, "Main Job", ""),
        ("2023-10-15", "Freelance", 3000, "Side Hustle", "")
    ]
    for row_data in dummy_income:
        ws_income.append(row_data)

    # --- Sheet 3: Expenses ---
    ws_expenses = wb.create_sheet("Expenses")
    headers_expenses = ["Date", "Description", "Amount", "Category", "Note"]

    for col_num, header in enumerate(headers_expenses, 1):
        cell = ws_expenses.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws_expenses.column_dimensions[get_column_letter(col_num)].width = 15

    # Add some dummy data
    dummy_expenses = [
        ("2023-10-02", "Rent", 6000, "Housing", ""),
        ("2023-10-05", "Groceries", 1500, "Food", ""),
        ("2023-10-10", "Internet", 599, "Utilities", "")
    ]
    for row_data in dummy_expenses:
        ws_expenses.append(row_data)

    # Save file
    output_path = "../templates/budget_template.xlsx"
    wb.save(output_path)
    print(f"Budget template created successfully at: {output_path}")

if __name__ == "__main__":
    create_budget_template()
